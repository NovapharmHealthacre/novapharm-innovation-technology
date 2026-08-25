from __future__ import annotations

import csv
import html
import os
import random
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import parse_qs, quote_plus, unquote, urljoin, urlparse

import requests
from openpyxl import load_workbook

BASE = Path("uk_pharmacy_sources")
OUT = BASE / "public_email_enrichment.csv"
SEARCH_URL = "https://html.duckduckgo.com/html/?q={}"
TIMEOUT = 20
MAX_WORKERS = 6

EMAIL_RX = re.compile(r"(?i)(?<![A-Z0-9._%+-])([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})(?![A-Z0-9._%+-])")
HREF_RX = re.compile(r'''(?i)href=["']([^"']+)["']''')
BAD_EMAIL_PARTS = {
    "example.com", "sentry.io", "wixpress.com", "cloudflare.com", "godaddy.com",
    "wordpress.org", "schema.org", "email.com", "domain.com", "yourdomain", "noreply",
}
BAD_DOMAINS = {
    "facebook.com", "instagram.com", "linkedin.com", "twitter.com", "x.com", "youtube.com",
    "yell.com", "yell.co.uk", "192.com", "cylex-uk.co.uk", "nhs.uk", "nhsinform.scot",
    "pharmdata.co.uk", "pharmanalyser.co.uk", "companieshouse.gov.uk", "gov.uk", "google.com",
    "bing.com", "duckduckgo.com", "trustpilot.com", "mapquest.com", "4ni.co.uk",
}

UA = "Mozilla/5.0 (compatible; NovaPharm public business contact research; +https://novapharmhealthcare.com)"


def clean_email(e: str) -> str | None:
    e = html.unescape(e).strip(" <>\"'()[]{}.,;:").lower()
    if len(e) > 120 or "@" not in e:
        return None
    if any(x in e for x in BAD_EMAIL_PARTS):
        return None
    local, domain = e.rsplit("@", 1)
    if not local or not domain or domain.endswith((".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp")):
        return None
    return e


def emails_from_text(text: str) -> list[str]:
    found = []
    for m in EMAIL_RX.finditer(html.unescape(text or "")):
        e = clean_email(m.group(1))
        if e and e not in found:
            found.append(e)
    return found


def resolve_ddg_link(href: str) -> str | None:
    href = html.unescape(href)
    if href.startswith("//"):
        href = "https:" + href
    if href.startswith("/"):
        u = urlparse(href)
        q = parse_qs(u.query)
        if "uddg" in q:
            return unquote(q["uddg"][0])
        return None
    if href.startswith("http://") or href.startswith("https://"):
        u = urlparse(href)
        if u.netloc.endswith("duckduckgo.com"):
            q = parse_qs(u.query)
            if "uddg" in q:
                return unquote(q["uddg"][0])
            return None
        return href
    return None


def usable_url(url: str) -> bool:
    try:
        host = urlparse(url).netloc.lower().split(":")[0]
    except Exception:
        return False
    if not host:
        return False
    return not any(host == d or host.endswith("." + d) for d in BAD_DOMAINS)


def fetch(session: requests.Session, url: str) -> tuple[str, str] | None:
    try:
        r = session.get(url, timeout=TIMEOUT, allow_redirects=True, headers={"User-Agent": UA})
        if r.status_code >= 400:
            return None
        ctype = (r.headers.get("content-type") or "").lower()
        if "text" not in ctype and "html" not in ctype and ctype:
            return None
        return r.url, r.text[:2_000_000]
    except Exception:
        return None


def discover_one(record: dict) -> dict:
    session = requests.Session()
    name = record["entity_name"]
    nation = record["nation"]
    postcode = record.get("postcode", "")
    query = f'"{name}" pharmacy {postcode} email contact UK'.strip()
    search_url = SEARCH_URL.format(quote_plus(query))
    out = dict(record)
    out.update({"query": query, "email": "", "email_source_url": "", "website": "", "confidence": "No public email found", "notes": ""})

    page = fetch(session, search_url)
    if not page:
        out["notes"] = "Search page unavailable"
        return out
    _, text = page

    # If the search result itself exposes a business email, keep it but mark as search-result evidence.
    result_emails = emails_from_text(text)
    if result_emails:
        out["email"] = result_emails[0]
        out["email_source_url"] = search_url
        out["confidence"] = "Public search-result email"

    links = []
    for h in HREF_RX.findall(text):
        u = resolve_ddg_link(h)
        if u and usable_url(u) and u not in links:
            links.append(u)
        if len(links) >= 5:
            break

    for candidate in links[:3]:
        got = fetch(session, candidate)
        if not got:
            continue
        final_url, body = got
        if usable_url(final_url) and not out["website"]:
            out["website"] = f"{urlparse(final_url).scheme}://{urlparse(final_url).netloc}/"
        ems = emails_from_text(body)
        if ems:
            out["email"] = ems[0]
            out["email_source_url"] = final_url
            out["confidence"] = "Published on business website"
            return out

        # Try obvious contact pages on the same business domain.
        root = f"{urlparse(final_url).scheme}://{urlparse(final_url).netloc}/"
        for path in ("contact", "contact-us", "contacts", "about/contact"):
            got2 = fetch(session, urljoin(root, path))
            if not got2:
                continue
            u2, body2 = got2
            ems2 = emails_from_text(body2)
            if ems2:
                out["email"] = ems2[0]
                out["email_source_url"] = u2
                out["website"] = root
                out["confidence"] = "Published on business contact page"
                return out

    return out


def build_entities() -> list[dict]:
    entities: dict[tuple[str, str], dict] = {}

    # Wales: contractor name is the legal/operator name.
    with (BASE / "nhsbsa_contractor_details_202607.csv").open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            if r.get("REGION_NAME", "") == "" and r.get("PHARMACY_TYPE") == "PHARMACY":
                name = (r.get("CONTRACTOR_NAME") or r.get("TRADING_NAME") or "").strip()
                if not name:
                    continue
                key = ("Wales", name.upper())
                entities.setdefault(key, {"nation": "Wales", "entity_name": name, "postcode": (r.get("POST_CODE") or "").strip(), "entity_scope": "Owner / contractor"})

    # Scotland: public file gives dispenser-location name; use it as the search entity.
    with (BASE / "scotland_dispenser_contactdetails_202604.csv").open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            name = (r.get("DispLocationName") or "").strip()
            if not name:
                continue
            key = ("Scotland", name.upper())
            entities.setdefault(key, {"nation": "Scotland", "entity_name": name, "postcode": (r.get("DispLocationPostcode") or "").strip(), "entity_scope": "Dispenser / operator name"})

    # Northern Ireland: NAME is the contractor/owner; T_A is trading name.
    wb = load_workbook(BASE / "northern_ireland_pharmaceutical_list_01.xlsx", read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(x or "") for x in next(rows)]
    for vals in rows:
        r = dict(zip(headers, vals))
        if str(r.get("CONTR_TYPE") or "").strip().lower() != "chemist":
            continue
        if str(r.get("OFF_LIST") or "").strip():
            continue
        name = str(r.get("NAME") or r.get("T_A") or "").strip()
        if not name:
            continue
        key = ("Northern Ireland", name.upper())
        entities.setdefault(key, {"nation": "Northern Ireland", "entity_name": name, "postcode": str(r.get("POSTCODE") or "").strip(), "entity_scope": "Owner / contractor"})

    return list(entities.values())


def main() -> None:
    entities = build_entities()
    print(f"Public-email enrichment entities: {len(entities)}")
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(discover_one, rec): rec for rec in entities}
        done = 0
        for fut in as_completed(futures):
            done += 1
            try:
                results.append(fut.result())
            except Exception as exc:
                rec = dict(futures[fut]); rec.update({"query":"","email":"","email_source_url":"","website":"","confidence":"Error","notes":repr(exc)})
                results.append(rec)
            if done % 100 == 0:
                print(f"Completed {done}/{len(entities)}")

    results.sort(key=lambda r: (r["nation"], r["entity_name"].upper()))
    fields = ["nation", "entity_name", "entity_scope", "postcode", "email", "website", "confidence", "email_source_url", "query", "notes"]
    with OUT.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader(); w.writerows(results)
    found = sum(bool(r.get("email")) for r in results)
    print(f"Published/search-result emails found: {found}/{len(results)}")


if __name__ == "__main__":
    main()
